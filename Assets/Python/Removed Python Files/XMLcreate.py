
import os
import numpy as np
import tensorflow as tf
import cv2
import xml.etree.ElementTree as ET
import matplotlib.pyplot as plt 
from collections import Counter 
from sklearn.cluster import KMeans 
import openpyxl
import pathlib
import math



def to_xml(windows_list, imdb_list, save_path, saveXML = False, printXML = False):

    """
    Creates an XML representation of window information.

    Parameters:
    - windows_list: List of window dictionaries
    - imdb_list: List of image database entries
    - save_path: Directory path to save the XML
    - saveXML: Flag to save the XML to a file (default: False)
    - printXML: Flag to print the XML content (default: False)
    """

    #for colour extraction testing
    #workbook = openpyxl.load_workbook(r'C:\Users\warre\Desktop\all_black_colours.xlsx')
    #worksheet1 = workbook.create_sheet("black")
    #worksheet2 = workbook.create_sheet("colour")

    root = ET.Element("AllWindows")
    #set model for window tpe detection
    current_directory = os.path.dirname(os.path.abspath(__file__))
    file_name = "windowTypeModel.keras"
    file_path = os.path.join(current_directory, file_name)
    model = tf.keras.models.load_model(file_path)  
    #iterate over all the images
    for s_idx in range(len(windows_list)):
        im = imdb_list[s_idx]['image']
        #set imagename in XML
        image_element = ET.Element("image")
        image_element.text = str(os.path.splitext(os.path.basename(im))[0])
        root.append(image_element)
        winPred = windows_list[s_idx]
        #set the colout of the facade into the xml
        black, kleur = get_colours(im, winPred, blur= False)
        color_elemenent = ET.Element("facade_color")
        image_element.append(color_elemenent)
        R_color = ET.SubElement(color_elemenent, "R")
        R_color.text = str(kleur[0])
        G_color = ET.SubElement(color_elemenent, "G")
        G_color.text = str(kleur[1])
        B_color = ET.SubElement(color_elemenent, "B")
        B_color.text = str(kleur[2])

        #for colour extraction testing
        #tup = tuple(black)
        #worksheet1.append(tup)
        #tup2 = tuple(kleur)
        #worksheet2.append(tup2)

        #sort detected windows by etage
        window = get_lowest_window(winPred)
        j = 0
        etages = {}
        while window:
            in_etage, out_etage = get_windows_in_etage(window, winPred)
            window = get_lowest_window(out_etage)
            winPred = out_etage
            etages["etage" + str(j)] = in_etage
            j += 1
        
        #set info of windows into XML per etage
        for etage in etages:
            etageList = etages[etage]
            etageElement = ET.Element(etage)
            image_element.append(etageElement)
            #way to set window info
            set_xml_average(etageElement, etageList, im, model)

    #for colour extraction testing
    #workbook.save(r'C:\Users\warre\Desktop\all_black_colours.xlsx')        
    xml1 = ET.ElementTree(root)   
    ET.indent(xml1)

    if printXML:
        print(ET.tostring(xml1.getroot(), encoding='unicode'))
    if saveXML:
        XMLfile = os.path.join(save_path, "result.xml")
        xml1.write(XMLfile)
        pathUnity = pathlib.Path(pathlib.Path(__file__).parent.parent.parent.parent.parent.resolve(),"Resources")
        XMLfile = os.path.join(pathUnity, "result.xml")
        xml1.write(XMLfile)
    
def get_image_size(img):
    """
    Gets the dimensions of an image.

    Parameters:
    - img: File path to the image or a NumPy array representing an image

    Returns:
    Tuple containing image dimensions (height, width, channels)
    """
    if isinstance(img, str):
        if os.path.exists(img):
            cv_img = cv2.imread(img)
    elif isinstance(img, np.ndarray):
        cv_img = img.copy()
    else:
        assert 0, "unKnown Type of img in XMLcreate get_size_image"
    return cv_img.shape

def get_lowest_window(windowList):
    """
    Gets the window with the lowest starting position. This is done using the mean of the vertical coordinates of each window.

    Parameters:
    - windowList: List of window dictionaries

    Returns:
    Window dictionary with the lowest position
    """
    if bool(windowList):
        window = windowList[0]
        for i in range(len(windowList)):
            new_window = np.array(windowList[i]['position'])[:, :2].copy()
            new_mean = np.mean(new_window, axis=0)[1]
            temp_window = np.array(window['position'])[:, :2].copy()
            temp_mean = np.mean(temp_window, axis=0)[1]
            if not bool(window):
                window = windowList[i]
            elif new_mean < temp_mean:
                window = windowList[i]
        return window
    else:
        return

def get_windows_in_etage(window, windowList):
    """
    Categorizes windows into in_etage and out_etage, meaning 'is situated inside the same etage or not'

    Parameters:
    - window: Reference window dictionary
    - windowList: List of window dictionaries

    Returns:
    Tuple containing lists of in_etage and out_etage window dictionaries
    """
    window_array = np.array(window['position'])[:, :2].copy()
    mean = np.mean(window_array, axis= 0)[1]
    max= np.max(window_array, axis=0)[1]
    min= np.min(window_array, axis=0)[1]
    size_win = max - min

    in_etage = []
    out_etage = []

    for i in range(len(windowList)):
        temp = np.array(windowList[i]['position'])[:, :2].copy()
        temp_mean = np.mean(temp, axis = 0)[1]
        if (mean - size_win/2)< temp_mean < (mean + size_win/2):
            in_etage.append(windowList[i])
        else:
            out_etage.append(windowList[i])
    return in_etage, out_etage

def set_xml_all_corners(etageElement, etageList, image):
    """
    Populates an etage XML element with window position information.
    Every window will be represented by all its corners.

    Parameters:
    - etageElement: XML element representing an etage
    - etageList: List of window dictionaries for the etage
    - image: File path to the image

    Returns:
    None
    """
    for i in range(len(etageList)):
        window_element = ET.Element("window")
        etageElement.append(window_element)
        windowName = ET.SubElement(window_element,'window_name')
        windowName.text = str(os.path.splitext(os.path.basename(image))[0]+ "_" + str(i+1))
        windowNumber = ET.SubElement(window_element, 'window_number')
        windowNumber.text = str(i+1)
        windowPosition = ET.SubElement(window_element, 'position')

        leftUpper = ET.SubElement(windowPosition, 'left_upper_corner')
        leftUpperWidth = ET.SubElement(leftUpper,'width')
        leftUpperWidth.text = str(np.array(etageList[i]['position'])[0, 0].copy())
        leftUpperHeight = ET.SubElement(leftUpper,'height')
        leftUpperHeight.text = str(np.array(etageList[i]['position'])[0, 1].copy())

        leftUnder = ET.SubElement(windowPosition, 'left_under_corner')
        leftUnderWidth = ET.SubElement(leftUnder,'width')
        leftUnderWidth.text = str(np.array(etageList[i]['position'])[1, 0].copy())
        leftUnderHeight = ET.SubElement(leftUnder,'height')
        leftUnderHeight.text = str(np.array(etageList[i]['position'])[1, 1].copy())

        rightUnder = ET.SubElement(windowPosition, 'right_under_corner')
        rightUnderWidth = ET.SubElement(rightUnder,'width')
        rightUnderWidth.text = str(np.array(etageList[i]['position'])[2, 0].copy())
        rightUnderHeight = ET.SubElement(rightUnder,'height')
        rightUnderHeight.text = str(np.array(etageList[i]['position'])[2, 1].copy())

        rightUpper = ET.SubElement(windowPosition, 'right_upper_corner')
        rightUpperWidth = ET.SubElement(rightUpper,'width')
        rightUpperWidth.text = str(np.array(etageList[i]['position'])[3, 0].copy())
        rightUpperHeight = ET.SubElement(rightUpper,'height')
        rightUpperHeight.text = str(np.array(etageList[i]['position'])[3, 1].copy())

def set_xml_pos_size(etageElement, etageList, image):
    """
    Populates an etage XML element with window position and size information.
    Position is the left under corner and the size is the average.

    Parameters:
    - etageElement: XML element representing an etage
    - etageList: List of window dictionaries for the etage
    - image: File path to the image

    Returns:
    None
    """
    im = cv2.imread(image)
    imHeight = im.shape
    for i in range(len(etageList)):
        window_element = ET.Element("window")
        etageElement.append(window_element)
        windowName = ET.SubElement(window_element,'window_name')
        windowName.text = str(os.path.splitext(os.path.basename(image))[0]+ "_" + str(i+1))
        windowNumber = ET.SubElement(window_element, 'window_number')
        windowNumber.text = str(i+1)
        windowPosition = ET.SubElement(window_element, 'position')

        averageWidth = np.mean(np.array(etageList[i]['position'])[2:4, 0].copy()) - np.mean(np.array(etageList[i]['position'])[0:2, 0].copy())
        averageHeight = np.mean(np.array(etageList[i]['position'])[1:3, 1].copy()) - ((np.array(etageList[i]['position'])[-1, 1].copy() + np.array(etageList[i]['position'])[0, 1].copy())/2)


        leftUnder = ET.SubElement(windowPosition, 'left_under_corner')
        leftUnderWidth = ET.SubElement(leftUnder,'width')
        leftUnderWidth.text = str(np.array(etageList[i]['position'])[1, 0].copy())
        leftUnderHeight = ET.SubElement(leftUnder,'height')
        leftUnderHeight.text = str(np.array(etageList[i]['position'])[1, 1].copy())

        windowSize = ET.SubElement(window_element, 'size')
        windowWidth = ET.SubElement(windowSize, 'width')
        windowWidth.text = str(averageWidth)
        windowHeight = ET.SubElement(windowSize, 'height')
        windowHeight.text = str(averageHeight)

def set_xml_relative(etageElement, etageList, image, model):
    """
    Populates an etage XML element with relative window position and size information.
    Relative in a way that the size of the image is 1-1.
    Every window gets its predicted windowtype.

    Parameters:
    - etageElement: XML element representing an etage
    - etageList: List of window dictionaries for the etage
    - image: File path to the image
    - model: Model for windowtype detection

    Returns:
    None
    """
    im = cv2.imread(image)
    imShape = im.shape
    imHeight = imShape[0]
    imWidth = imShape[1]
    for i in range(len(etageList)):
        
        window_element = ET.Element("window")
        etageElement.append(window_element)
        windowName = ET.SubElement(window_element,'window_name')
        windowName.text = str(os.path.splitext(os.path.basename(image))[0]+ "_" + str(i+1))
        windowNumber = ET.SubElement(window_element, 'window_number')
        windowNumber.text = str(i+1)
        windowPosition = ET.SubElement(window_element, 'position')

        averageWidth = (np.mean(np.array(etageList[i]['position'])[2:4, 0].copy()) - np.mean(np.array(etageList[i]['position'])[0:2, 0].copy()))/imWidth
        averageHeight = (np.mean(np.array(etageList[i]['position'])[1:3, 1].copy()) - ((np.array(etageList[i]['position'])[-1, 1].copy() + np.array(etageList[i]['position'])[0, 1].copy())/2))/imHeight

        leftUnder = ET.SubElement(windowPosition, 'left_under_corner')
        leftUnderWidth = ET.SubElement(leftUnder,'width')
        leftUnderWidth.text = str((np.array(etageList[i]['position'])[1, 0].copy())/imWidth)
        leftUnderHeight = ET.SubElement(leftUnder,'height')
        leftUnderHeight.text = str((imHeight - np.array(etageList[i]['position'])[1, 1].copy())/imHeight)

        windowSize = ET.SubElement(window_element, 'size')
        windowWidth = ET.SubElement(windowSize, 'width')
        windowWidth.text = str(averageWidth)
        windowHeight = ET.SubElement(windowSize, 'height')
        windowHeight.text = str(averageHeight)

        windowType = ET.SubElement(window_element, 'type')
        rectWin = get_window_image(etageList[i], im)
        classOfWindow = predict_windowTpye(rectWin, model)
        windowType.text = classOfWindow

def set_xml_average(etageElement, etageList, image, model):
    """
    Populates an etage XML element with relative window position and size information.
    Relative in a way that the size of the image is 1-1.
    Every window gets its predicted windowtype.

    Parameters:
    - etageElement: XML element representing an etage
    - etageList: List of window dictionaries for the etage
    - image: File path to the image
    - model: Model for windowtype detection

    Returns:
    None
    """
    im = cv2.imread(image)
    imShape = im.shape
    imHeight = imShape[0]
    imWidth = imShape[1]
    imageSize = get_image_size(im)

    threshold = 0.02 * imageSize[0]

    groups1 = group_windows_by_y_position(etageList, threshold)
    

    for l in range(len(groups1)):
        group1 = groups1[l]
        average_position = calculate_average_group_position(group1)
        groups2 = group_windows_by_size(group1, threshold, False)
        for k in range(len(groups2)):
            group2 = groups2[k]
            average_x_size = calculate_average_group_size(group2, False)
            groups3 = group_windows_by_size(group2, threshold, True)
            for j in range(len(groups3)):
                group3 = groups3[j]
                average_y_size = calculate_average_group_size(group3, True)
                for i in range(len(group3)):
                    window_element = ET.Element("window")
                    etageElement.append(window_element)
                    windowName = ET.SubElement(window_element,'window_name')
                    windowName.text = str(os.path.splitext(os.path.basename(image))[0]+ "_" + str(i+1))
                    windowNumber = ET.SubElement(window_element, 'window_number')
                    windowNumber.text = str(i+1)
                    windowPosition = ET.SubElement(window_element, 'position')

                    averageWidth = (average_x_size)/imWidth
                    averageHeight = (average_y_size)/imHeight

                    leftUnder = ET.SubElement(windowPosition, 'left_under_corner')
                    leftUnderWidth = ET.SubElement(leftUnder,'width')
                    leftUnderWidth.text = str((np.array(group3[i]['position'])[1, 0].copy())/imWidth)
                    leftUnderHeight = ET.SubElement(leftUnder,'height')
                    leftUnderHeight.text = str((imHeight - average_position)/imHeight)

                    windowSize = ET.SubElement(window_element, 'size')
                    windowWidth = ET.SubElement(windowSize, 'width')
                    windowWidth.text = str(averageWidth)
                    windowHeight = ET.SubElement(windowSize, 'height')
                    windowHeight.text = str(averageHeight)

                    windowType = ET.SubElement(window_element, 'type')
                    rectWin = get_window_image(group3[i], im, True)
                    classOfWindow = predict_windowTpye(rectWin, model)
                    windowType.text = classOfWindow      

def calculate_window_y_position(window):
    window_height = (np.array(window['position'])[2, 1] + np.array(window['position'])[1, 1])/2
    return window_height

def calculate_windows_size(window, vertical = False):
    """
    Calculate the size of a rectangular window based on its corner coordinates.
    
    Parameters:
    window (dict): A dictionary containing the corner coordinates of the window.
    
    Returns:
    float: The size of the window calculated.
    """
    if vertical:
        size =  (np.mean(np.array(window['position'])[1:3, 1].copy()) - ((np.array(window['position'])[-1, 1].copy() + np.array(window['position'])[0, 1].copy())/2))
    else:
        size = (np.mean(np.array(window['position'])[2:4, 0].copy()) - np.mean(np.array(window['position'])[0:2, 0].copy()))
    # square1 = (window['position'])[1, 0] - window['position'])[3, 0])**2
    # square2 = (window['position'])[1, 1] - window['position'])[3, 1])**2
    # diagonal = math.sqrt(square1 + square2)
    return size

def calculate_average_group_position(data_list):
    """
    Calculate the average y-coordinate value of a list of data items, each containing a 'position' array.

    Parameters:
    data_list (list): A list of dictionaries, where each dictionary contains a 'position' array.
    
    Returns:
    float: The average y-coordinate value calculated from the 'position' arrays in the data list.
           If the list is empty, the function returns 0 to avoid division by zero.
    """
    total_sum = 0
    num_elements = len(data_list)

    for item in data_list:
        position_array = item['position']
        y_value = position_array[1, 1]  # Extract [1, 1] element from the position array
        total_sum += y_value

    if num_elements > 0:
        average_position = total_sum / num_elements
    else:
        average_position = 0  # Avoid division by zero if the list is empty
    
    return average_position

def calculate_average_group_size(data_list, vertical = False):
    """
    Calculate the average y-coordinate value of a list of data items, each containing a 'position' array.

    Parameters:
    data_list (list): A list of dictionaries, where each dictionary contains a 'position' array.
    
    Returns:
    float: The average y-coordinate value calculated from the 'position' arrays in the data list.
           If the list is empty, the function returns 0 to avoid division by zero.
    """
    total_sum = 0
    num_elements = len(data_list)

    for item in data_list:
        size_value = calculate_windows_size(item, vertical)
        total_sum += size_value

    if num_elements > 0:
        average_size = total_sum / num_elements
    else:
        average_size = 0  # Avoid division by zero if the list is empty
    
    return average_size

def group_windows_by_y_position(windowList, threshold):
    groups = []
    
    while windowList:
        current_window = windowList.pop(0)
        current_group = [current_window]
        i = 0

        while i < len(windowList):
            diff = abs((calculate_window_y_position(windowList[i])) - (calculate_window_y_position(current_window)))

            if diff <= threshold:
                current_group.append(windowList.pop(i))
            else:
                i += 1
        
        groups.append(current_group)

    return groups

def group_windows_by_size(windowList, threshold, vertical = False):
    groups = []
    
    while windowList:
        current_window = windowList.pop(0)
        current_group = [current_window]
        i = 0

        while i < len(windowList):
            diff = abs(calculate_windows_size(windowList[i], vertical) -  calculate_windows_size(current_window, vertical))

            if diff <= threshold:
                current_group.append(windowList.pop(i))
            else:
                i += 1
        
        groups.append(current_group)

    return groups

def get_window_image(window, img, expand = False):
    """
    get a rectified image of a window from an image.

    Parameters:
    - window: Reference window dictionary
    - image: Original image

    Returns:
    Rectified window image
    """
    imageSize = get_image_size(img)
    
    if expand:
        threshold_y = 0.01 * imageSize[0]
        threshold_x = 0.02 * imageSize[1]
    else:
        threshold_y = 0 
        threshold_x = 0 
    # Rectification
    rows,cols,ch = img.shape
    pts1 = np.float32([[np.array(window['position'])[0, 0].copy(), np.array(window['position'])[0, 1].copy()-threshold_y],
                    [np.array(window['position'])[3, 0].copy(), np.array(window['position'])[3, 1].copy()-threshold_y],
                    [np.array(window['position'])[2, 0].copy(), np.array(window['position'])[2, 1].copy()],
                    [np.array(window['position'])[1, 0].copy(), np.array(window['position'])[1, 1].copy()]])
    
    #sorting all cornors from left up, right up, right down to left down
    sorted_indices = np.argsort(pts1[:, 1])
    sorted_pts1 = pts1[sorted_indices]
    if(sorted_pts1[0,0]> sorted_pts1[1,0]):
        sorted_pts1[[0,1]] = sorted_pts1[[1,0]]
        

    if(sorted_pts1[2,0]< sorted_pts1[3,0]):
        sorted_pts1[[2,3]] = sorted_pts1[[3,2]]
        
    #get the average width and height that was selected to use as dimensions of resulting image
    mean_x = int(round(((sorted_pts1[1,0]+sorted_pts1[2,0])-(sorted_pts1[0,0]+sorted_pts1[3,0]))/2, 0))
    mean_y = int(round(((sorted_pts1[3,1]+sorted_pts1[2,1])-(sorted_pts1[0,1]+sorted_pts1[1,1]))/2, 0))
    pts2 = np.float32([[0,0],[mean_x,0],[mean_x,mean_y],[0,mean_y]])
    M = cv2.getPerspectiveTransform(sorted_pts1,pts2)
    dst = cv2.warpPerspective(img,M,(mean_x,mean_y))
    # dst2 = cv2.cvtColor(dst, cv2.COLOR_RGB2BGR)
    # cv2.imshow('image', dst2)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    return dst

def predict_windowTpye(image, model):
    class_indices = {'class_1': 0, 'class_2': 1, 'class_3': 2, 'class_4': 3, 'class_5': 4}

    target_size = (224, 224)
    img_resized = cv2.resize(image, target_size)

    # Convert the image to a NumPy array
    img_array = np.expand_dims(img_resized, axis=0)

    # Preprocess the image using MobileNetV2 preprocessing
    img_array = tf.keras.applications.mobilenet_v2.preprocess_input(img_array)

    predictions = model.predict(img_array)
    print(predictions)
    predicted_class_index = np.argmax(predictions)
    certainty = predictions[0][predicted_class_index]
    print(certainty)
    predicted_class_name = [k for k, v in class_indices.items() if v == predicted_class_index][0]
    return predicted_class_name

def filter_neutral_colors(img_og, lower = (0, 0, 0), upper = (500, 0,255)):
    """
    Filters out neutral colors from an image.

    Parameters:
    - img_og: Original image
    - lower: Lower HSV threshold for neutral colors (default: (0, 0, 0))
    - upper: Upper HSV threshold for neutral colors (default: (500, 0, 255))

    Returns:
    Filtered image
    """    
    # convert to hsv
    hsv = cv2.cvtColor(img_og, cv2.COLOR_BGR2HSV)
    # mask of gray (0,0,0) ~ (0, 0,255)
    mask = cv2.inRange(hsv, lower, upper)
    # filter out gray colours (convert them to white, which will be excluded at the end)
    img_filtered = cv2.bitwise_and(img_og,img_og, mask=cv2.bitwise_not(mask))
    # print("Original sum check: ",np.sum(img_og))
    # print("Filtered sum check: ",np.sum(img_filtered))
    return img_filtered

def remove_windows(image, windowList, expand = False):
    """
    Removes windows from an image. Turns it black.

    Parameters:
    - image: Original image
    - windowList: List of window dictionaries
    - expand: Flag to expand removal area (default: False)

    Returns:
    Image with windows removed
    """
    mask = np.zeros(image.shape[:2], dtype="uint8")
    for i in range(len(windowList)):
        if expand:
            cv2.rectangle(mask, (int(np.array(windowList[i]['position'])[0, 0].copy())-10,int(np.array(windowList[i]['position'])[0, 1].copy())-10), 
                (int(np.array(windowList[i]['position'])[2, 0].copy())+10,int(np.array(windowList[i]['position'])[2, 1].copy())+10), 255, -1)
        else:
            cv2.rectangle(mask, (int(np.array(windowList[i]['position'])[0, 0].copy()),int(np.array(windowList[i]['position'])[0, 1].copy())), 
                (int(np.array(windowList[i]['position'])[2, 0].copy()),int(np.array(windowList[i]['position'])[2, 1].copy())), 255, -1)  
    masked = cv2.bitwise_and(image, image, mask=cv2.bitwise_not(mask))
    return masked

def get_colours(image, windowList, blur = False):
    """
    Performs color analysis on an image of a facade.
    Removes the windows, turn them black and get the most domninant colours in the image. This will be the black and the facade colout. 
    Than assert wich one is the black and which one is the colour.

    Parameters:
    - image: Image file path
    - windowList: List of window dictionaries
    - blur: Flag to apply Gaussian blur to the image (default: False)

    Returns:
    Black color (window frames) and Color color (facade elements)
    """
    img = cv2.imread(image)
    if blur:
        img = cv2.GaussianBlur(img, (33,21),0)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    #filter out the blacks and greys
    img_filtered = filter_neutral_colors(img, lower = (0, 0, 0), upper = (180, 255, 40))
    #filter out the over saturated parts
    img_filtered = filter_neutral_colors(img_filtered, lower = (0, 0, 0), upper = (180, 10, 255))
    img_filtered = remove_windows(img_filtered, windowList, expand=True)
    plt.imshow(img_filtered)
    #plt.show()

    img_filtered = img_filtered.reshape((img_filtered.shape[0] * img_filtered.shape[1],3)) #represent as row*column,channel number
    clf = KMeans(n_clusters =2, n_init= 'auto')
    color_labels = clf.fit_predict(img_filtered)
    center_colors = clf.cluster_centers_
    counts = Counter(color_labels)
    ordered_colors = []
    values = []
    black = []
    kleur = []
    for k in sorted(counts, key=counts.get, reverse= True):
        colour = center_colors[k]
        ordered_colors.append(center_colors[k])
        values.append(counts[k])
    
    color1 = ordered_colors[0]
    color2 = ordered_colors[1]
    if(np.mean(color1)<np.mean(color2)):
        kleur = color2
        black = color1
    else:
        kleur = color1
        black = color2

    hex_colors = [rgb_to_hex(ordered_colors[i]) for i in range(len(ordered_colors))]

    plt.figure(figsize = (12, 8))
    plt.pie(values, labels = hex_colors, colors = hex_colors)

    #plt.show()
    #plt.savefig("results/my_pie.png")

    # print("Found the following colors:\n")
    # for color in hex_colors:
    #     print(color)
    return black, kleur

def rgb_to_hex(rgb_color):
    """
    Converts an RGB color tuple to a hexadecimal representation.

    Parameters:
    - rgb_color: Tuple representing an RGB color

    Returns:
    Hexadecimal color representation
    """
    hex_color = "#"
    for i in rgb_color:
        hex_color += ("{:02x}".format(int(i)))
    return hex_color



